import openpyxl
from openpyxl.utils import get_column_letter
import tempfile
from openpyxl.styles import Font, Border, Side, Alignment

class ExcelProcessor:
    @staticmethod
    async def create_excel_file(data):
        # Create a new Excel workbook and activate the worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        container_name=data['container_name']
        # Define styles
        times_new_roman_14 = Font(name='Times New Roman', size=14)  # Общий стиль шрифта
        bold_font = Font(name='Times New Roman', size=14, bold=True)  # Жирный стиль
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))  # Границы ячеек
        center_alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание по центру

        # Set a standard column width
        for col in range(1, 20):  # На случай большого количества данных (1-20 колонок)
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 50

        # Fill container information with bold text, custom font, and borders
        ws['A1'] = "Container"
        ws['A1'].font = bold_font
        ws['A1'].border = thin_border
        ws['A1'].alignment = center_alignment

        ws['B1'] = f"{data['container_name']}"
        ws['B1'].font = bold_font
        ws['B1'].border = thin_border
        ws['B1'].alignment = center_alignment

        # Initialize the start row and column
        row_start = 3
        col_start = 1

        # Write headers for each article and apply bold font, custom font, borders, and center alignment
        # Write headers for each article and apply bold font, custom font, borders, and center alignment
        for scanned_item in data.get('scanned', []):
            article = scanned_item['article']
            count = len(scanned_item['dms'])  # Подсчитываем количество dm_without_tail
            header_text = f"{article} ( {count} pcs )"
            column_letter = get_column_letter(col_start)

            ws[f"{column_letter}{row_start}"] = header_text
            ws[f"{column_letter}{row_start}"].font = bold_font
            ws[f"{column_letter}{row_start}"].border = thin_border
            ws[f"{column_letter}{row_start}"].alignment = center_alignment

            # Write all dm_without_tail values under the article with borders, font style, and center alignment
            row_dm = row_start + 1
            for dm in scanned_item['dms']:
                ws[f"{column_letter}{row_dm}"] = dm['dm_without_tail']
                ws[f"{column_letter}{row_dm}"].font = times_new_roman_14
                ws[f"{column_letter}{row_dm}"].border = thin_border
                ws[f"{column_letter}{row_dm}"].alignment = center_alignment
                row_dm += 1

            # Move to the next column for the next article
            col_start += 1


        # Save the Excel workbook to a temporary file
        file_name = f"{container_name}_kit"
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=file_name)
        wb.save(temp_file.name)

        return temp_file.name, temp_file.name.split("\\")[-1]  # Return file path and name


    @staticmethod
    async def create_excel_file(data):
        """
        Текущий формат (как сейчас):
        - 1 контейнер
        - артикулы = колонки
        - под каждым артикулом список dm_without_tail
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        container_name = data["container_name"]

        times_new_roman_14 = Font(name="Times New Roman", size=14)
        bold_font = Font(name="Times New Roman", size=14, bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # ширина колонок "с запасом"
        for col in range(1, 20):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 50

        ws["A1"] = "Container"
        ws["A1"].font = bold_font
        ws["A1"].border = thin_border
        ws["A1"].alignment = center_alignment

        ws["B1"] = f"{data['container_name']}"
        ws["B1"].font = bold_font
        ws["B1"].border = thin_border
        ws["B1"].alignment = center_alignment

        row_start = 3
        col_start = 1

        for scanned_item in data.get("scanned", []):
            article = scanned_item["article"]
            count = len(scanned_item["dms"])
            header_text = f"{article} ( {count} pcs )"

            column_letter = get_column_letter(col_start)
            ws[f"{column_letter}{row_start}"] = header_text
            ws[f"{column_letter}{row_start}"].font = bold_font
            ws[f"{column_letter}{row_start}"].border = thin_border
            ws[f"{column_letter}{row_start}"].alignment = center_alignment

            row_dm = row_start + 1
            for dm in scanned_item["dms"]:
                ws[f"{column_letter}{row_dm}"] = dm["dm_without_tail"]
                ws[f"{column_letter}{row_dm}"].font = times_new_roman_14
                ws[f"{column_letter}{row_dm}"].border = thin_border
                ws[f"{column_letter}{row_dm}"].alignment = center_alignment
                row_dm += 1

            col_start += 1

        file_prefix = f"{container_name}_kit"
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=file_prefix)
        wb.save(temp_file.name)

        # Windows-friendly name
        return temp_file.name, temp_file.name.split("\\")[-1]

    @staticmethod
    async def create_excel_file_bulk_single_sheet(
            kits: list[dict],
            file_prefix: str = "containers_bulk_",
            *,
            order_by: str = "container_id",  # container_id | packed_date | container_name
            order_dir: str = "asc",  # asc | desc
            sort_articles: str = "asc",  # asc | desc
            sort_dms: str = "asc",  # asc | desc
    ):
        """
        Новый формат:
        - 1 Excel файл
        - 1 лист
        - 1 таблица (строки)
        Одна строка = один DM-код.

        Колонки:
        packed_date | container_id | container_name | article | dm_without_tail | article_qty
        """

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Containers"

        font = Font(name="Times New Roman", size=14)
        bold = Font(name="Times New Roman", size=14, bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center")

        headers = ["packed_date", "container_id", "container_name", "article", "dm_without_tail", "article_qty"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.border = thin_border
            cell.alignment = center

        # фильтры/фикс заголовка
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        widths = [25, 14, 28, 18, 50, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # сортировка контейнеров внутри Excel (на всякий случай)
        reverse = (order_dir.lower() == "desc")

        def key_fn(k: dict):
            if order_by == "packed_date":
                return k.get("packed_date") or ""
            if order_by == "container_name":
                return k.get("container_name") or ""
            return k.get("container_id") or 0

        kits_sorted = sorted(kits, key=key_fn, reverse=reverse)

        row = 2
        for kit in kits_sorted:
            packed_date = kit.get("packed_date") or ""
            cid = kit.get("container_id") or ""
            cname = kit.get("container_name") or ""

            scanned = kit.get("scanned", []) or []
            scanned_sorted = sorted(
                scanned,
                key=lambda x: x.get("article") or "",
                reverse=(sort_articles.lower() == "desc"),
            )

            # если контейнер пустой — оставим 1 строку (чтобы видно было, что он вошел)
            if not scanned_sorted:
                values = [packed_date, cid, cname, "", "", 0]
                for col, v in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = font
                    cell.border = thin_border
                    cell.alignment = center
                row += 1
                continue

            for item in scanned_sorted:
                article = item.get("article") or ""
                dms = item.get("dms", []) or []
                qty = len(dms)

                dms_sorted = sorted(
                    dms,
                    key=lambda d: d.get("dm_without_tail") or "",
                    reverse=(sort_dms.lower() == "desc"),
                )

                for dm in dms_sorted:
                    values = [packed_date, cid, cname, article, dm.get("dm_without_tail") or "", qty]
                    for col, v in enumerate(values, start=1):
                        cell = ws.cell(row=row, column=col, value=v)
                        cell.font = font
                        cell.border = thin_border
                        cell.alignment = center
                    row += 1

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=file_prefix)
        wb.save(temp_file.name)

        return temp_file.name, temp_file.name.split("\\")[-1]
